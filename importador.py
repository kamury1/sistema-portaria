import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

import openpyxl

from banco import cadastrar_morador
from banco import listar_moradores

from banco import cadastrar_prestador
from banco import listar_prestadores


# ==========================================================
# NORMALIZAR TEXTO
# ==========================================================

def normalizar(valor):

    if valor is None:
        return ""

    return str(valor).strip().lower()


# ==========================================================
# CONVERTER VALOR PARA TEXTO
# ==========================================================

def texto(valor):

    if valor is None:
        return ""

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    return str(valor).strip()


# ==========================================================
# IDENTIFICAR TIPO DO MORADOR
# ==========================================================

def identificar_tipo(observacao):

    obs = normalizar(observacao)

    if "propriet" in obs:
        return "Proprietário"

    if "locat" in obs:
        return "Inquilino"

    if "inquilin" in obs:
        return "Inquilino"

    return "Outro"


# ==========================================================
# LOCALIZAR ABA
# ==========================================================

def localizar_aba(workbook, nome_procurado):

    nome_procurado = (
        nome_procurado
        .lower()
        .replace("_", " ")
        .strip()
    )

    for aba in workbook.sheetnames:

        nome_aba = (
            aba
            .lower()
            .replace("_", " ")
            .strip()
        )

        if nome_aba == nome_procurado:
            return aba

    return None


# ==========================================================
# IMPORTAR MORADORES
# ==========================================================

def importar_moradores(workbook):

    print("\n==========================================")
    print("ABAS ENCONTRADAS NO EXCEL:")
    print("==========================================")

    for aba in workbook.sheetnames:
        print(repr(aba))

    print("==========================================\n")

    nome_aba = localizar_aba(
        workbook,
        "Cadastro Moradores"
    )

    print(
        "ABA DE MORADORES LOCALIZADA:",
        repr(nome_aba)
    )

    if nome_aba is None:

        print(
            "ERRO: Aba Cadastro Moradores não encontrada."
        )

        return 0, 0, 0

    planilha = workbook[nome_aba]

    # ======================================================
    # TESTE DA ABA DE MORADORES
    # ======================================================

    print(
        "LINHAS NA ABA MORADORES:",
        planilha.max_row
    )

    print(
        "COLUNAS NA ABA MORADORES:",
        planilha.max_column
    )

    print("\nPRIMEIRAS LINHAS:")

    for numero, linha in enumerate(
        planilha.iter_rows(
            min_row=1,
            max_row=5,
            values_only=True
        ),
        start=1
    ):

        print(
            "LINHA",
            numero,
            "=",
            linha
        )

    print("\n")

    print(
        "LENDO MORADORES DA ABA:",
        repr(nome_aba)
    )

    moradores_existentes = listar_moradores()

    registros_existentes = set()

    for morador in moradores_existentes:

        nome = normalizar(
            morador[1]
        )

        apartamento = normalizar(
            morador[2]
        )

        registros_existentes.add(
            (
                nome,
                apartamento
            )
        )

    importados = 0
    duplicados = 0
    ignorados = 0

    # ------------------------------------------------------
    # ESTRUTURA DA PLANILHA
    #
    # A = Código
    # B = Nome
    # C = Unidade
    # D = Telefone
    # E = Veículo
    # F = Vagas
    # G = Observações
    # ------------------------------------------------------

    for linha in planilha.iter_rows(
        min_row=3,
        values_only=True
    ):

        if len(linha) < 7:

            ignorados += 1
            continue

        nome = texto(
            linha[1]
        )

        apartamento = texto(
            linha[2]
        )

        telefone = texto(
            linha[3]
        )

        vagas = texto(
            linha[5]
        )

        observacao = texto(
            linha[6]
        )

        if nome == "":
            continue

        if apartamento == "":

            ignorados += 1
            continue

        chave = (
            normalizar(nome),
            normalizar(apartamento)
        )

        if chave in registros_existentes:

            duplicados += 1
            continue

        tipo = identificar_tipo(
            observacao
        )

        cadastrar_morador(
            nome,
            apartamento,
            vagas,
            telefone,
            tipo
        )

        registros_existentes.add(
            chave
        )

        importados += 1

    print("------------------------------------------")
    print("RESULTADO MORADORES")
    print("Importados:", importados)
    print("Duplicados:", duplicados)
    print("Ignorados:", ignorados)
    print("------------------------------------------")

    return (
        importados,
        duplicados,
        ignorados
    )


# ==========================================================
# IMPORTAR PRESTADORES
# ==========================================================

def importar_prestadores(workbook):

    nome_aba = localizar_aba(
        workbook,
        "Cadastro Prestadores"
    )

    if nome_aba is None:
        return 0, 0, 0

    planilha = workbook[nome_aba]

    prestadores_existentes = listar_prestadores()

    registros_existentes = set()

    for prestador in prestadores_existentes:

        nome = normalizar(
            prestador[1]
        )

        documento = normalizar(
            prestador[2]
        )

        registros_existentes.add(
            (
                nome,
                documento
            )
        )

    importados = 0
    duplicados = 0
    ignorados = 0

    # ------------------------------------------------------
    # ESTRUTURA DA PLANILHA
    #
    # A = Código
    # B = Nome
    # C = Documento
    # D = Empresa
    # E = Serviço
    # F = Telefone
    # G = Observações
    # ------------------------------------------------------

    for linha in planilha.iter_rows(
        min_row=3,
        values_only=True
    ):

        if len(linha) < 6:

            ignorados += 1
            continue

        nome = texto(
            linha[1]
        )

        documento = texto(
            linha[2]
        )

        empresa = texto(
            linha[3]
        )

        telefone = texto(
            linha[5]
        )

        if nome == "":
            continue

        if documento == "":

            ignorados += 1
            continue

        chave = (
            normalizar(nome),
            normalizar(documento)
        )

        if chave in registros_existentes:

            duplicados += 1
            continue

        cadastrar_prestador(
            nome,
            documento,
            empresa,
            telefone
        )

        registros_existentes.add(
            chave
        )

        importados += 1

    return (
        importados,
        duplicados,
        ignorados
    )


# ==========================================================
# SELECIONAR E IMPORTAR PLANILHA
# ==========================================================

def importar_planilha(janela_principal):

    caminho = filedialog.askopenfilename(
        parent=janela_principal,
        title="Selecionar planilha para importar",
        filetypes=[
            (
                "Planilhas do Excel",
                "*.xlsx *.xlsm"
            ),
            (
                "Todos os arquivos",
                "*.*"
            )
        ]
    )

    if caminho == "":
        return

    confirmar = messagebox.askyesno(
        "Importar dados",
        "Deseja importar os moradores e prestadores "
        "desta planilha?\n\n"
        "Registros já cadastrados serão ignorados.",
        parent=janela_principal
    )

    if not confirmar:
        return

    workbook = None

    try:

        workbook = openpyxl.load_workbook(
            caminho,
            data_only=True,
            read_only=True
        )

        (
            moradores_importados,
            moradores_duplicados,
            moradores_ignorados
        ) = importar_moradores(
            workbook
        )

        (
            prestadores_importados,
            prestadores_duplicados,
            prestadores_ignorados
        ) = importar_prestadores(
            workbook
        )

        workbook.close()
        workbook = None

        mensagem = (
            "IMPORTAÇÃO CONCLUÍDA!\n\n"

            "MORADORES\n"
            f"Importados: {moradores_importados}\n"
            f"Já cadastrados: {moradores_duplicados}\n"
            f"Ignorados: {moradores_ignorados}\n\n"

            "PRESTADORES\n"
            f"Importados: {prestadores_importados}\n"
            f"Já cadastrados: {prestadores_duplicados}\n"
            f"Ignorados: {prestadores_ignorados}"
        )

        messagebox.showinfo(
            "Importação concluída",
            mensagem,
            parent=janela_principal
        )

    except PermissionError:

        messagebox.showerror(
            "Erro",
            "Não foi possível abrir a planilha.\n\n"
            "Feche o arquivo no Excel e tente novamente.",
            parent=janela_principal
        )

    except Exception as erro:

        print("\nERRO NA IMPORTAÇÃO:")
        print(repr(erro))

        messagebox.showerror(
            "Erro na importação",
            f"Não foi possível importar a planilha.\n\n"
            f"Erro: {erro}",
            parent=janela_principal
        )

    finally:

        if workbook is not None:

            try:
                workbook.close()

            except Exception:
                pass