import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from banco import (
    listar_prestadores,
    registrar_entrada_prestador,
    listar_acessos_ativos,
    registrar_saida_prestador,
    listar_historico_acessos,
)

COR_FUNDO = "#0B1220"
COR_TOPO = "#111D30"
COR_CARTAO = "#17263A"
COR_BORDA = "#253B56"
COR_CAMPO = "#0F1B2D"
COR_BOTAO = "#1B2D45"
COR_BOTAO_HOVER = "#274463"
COR_SUCESSO = "#2E7D32"
COR_SUCESSO_HOVER = "#256428"
COR_PERIGO = "#C62828"
COR_PERIGO_HOVER = "#A51F1F"
COR_TEXTO = "#F5F7FA"
COR_TEXTO_2 = "#A9B7C6"
COR_TABELA = "#FFFFFF"
COR_TABELA_ALT = "#F3F6FA"
COR_TABELA_TEXTO = "#17202A"
COR_SELECAO = "#D9E8F6"
FONTE = "Segoe UI"


def centralizar(janela, largura, altura):
    janela.update_idletasks()
    x = max((janela.winfo_screenwidth() - largura) // 2, 0)
    y = max((janela.winfo_screenheight() - altura) // 2, 0)
    janela.geometry(f"{largura}x{altura}+{x}+{y}")


def hover(botao_widget, normal, sobre):
    botao_widget.bind("<Enter>", lambda e: botao_widget.config(bg=sobre))
    botao_widget.bind("<Leave>", lambda e: botao_widget.config(bg=normal))


def criar_botao(parent, texto, comando, cor=COR_BOTAO, sobre=COR_BOTAO_HOVER):
    b = tk.Button(
        parent, text=texto, command=comando,
        font=(FONTE, 10, "bold"),
        bg=cor, fg=COR_TEXTO,
        activebackground=sobre, activeforeground=COR_TEXTO,
        relief="flat", bd=0, padx=18, pady=9, cursor="hand2"
    )
    hover(b, cor, sobre)
    return b


def criar_entry(parent, largura=None):
    return tk.Entry(
        parent, width=largura, font=(FONTE, 11),
        bg=COR_CAMPO, fg=COR_TEXTO, insertbackground=COR_TEXTO,
        relief="flat", bd=0, highlightthickness=1,
        highlightbackground=COR_BORDA, highlightcolor="#4C78A8"
    )


def configurar_tabela(estilo_nome):
    estilo = ttk.Style()
    try:
        estilo.theme_use("clam")
    except tk.TclError:
        pass

    estilo.configure(
        estilo_nome,
        background=COR_TABELA,
        fieldbackground=COR_TABELA,
        foreground=COR_TABELA_TEXTO,
        rowheight=34,
        borderwidth=0,
        font=(FONTE, 10),
    )
    estilo.map(
        estilo_nome,
        background=[("selected", COR_SELECAO)],
        foreground=[("selected", COR_TABELA_TEXTO)],
    )
    estilo.configure(
        estilo_nome + ".Heading",
        background=COR_TOPO,
        foreground=COR_TEXTO,
        relief="flat",
        padding=(8, 10),
        font=(FONTE, 10, "bold"),
    )


def criar_topo(janela, titulo, subtitulo):
    frame = tk.Frame(janela, bg=COR_TOPO, height=96)
    frame.pack(fill="x")
    frame.pack_propagate(False)

    tk.Label(
        frame, text=titulo, font=(FONTE, 20, "bold"),
        bg=COR_TOPO, fg=COR_TEXTO
    ).pack(anchor="w", padx=30, pady=(18, 0))

    tk.Label(
        frame, text=subtitulo, font=(FONTE, 9),
        bg=COR_TOPO, fg=COR_TEXTO_2
    ).pack(anchor="w", padx=31, pady=(3, 0))


def abrir_registro_entrada(janela_principal, operador_atual):
    janela = tk.Toplevel(janela_principal)
    janela.title("Registrar Entrada de Prestador")
    janela.configure(bg=COR_FUNDO)
    janela.resizable(False, False)
    centralizar(janela, 760, 500)
    janela.transient(janela_principal)

    criar_topo(
        janela,
        "Registrar Entrada",
        "Libere a entrada de um prestador previamente cadastrado."
    )

    card = tk.Frame(
        janela, bg=COR_CARTAO,
        highlightthickness=1, highlightbackground=COR_BORDA
    )
    card.pack(fill="both", expand=True, padx=30, pady=26)

    form = tk.Frame(card, bg=COR_CARTAO)
    form.pack(fill="x", padx=32, pady=(30, 10))
    form.grid_columnconfigure(1, weight=1)

    tk.Label(
        form, text="Prestador *", font=(FONTE, 10, "bold"),
        bg=COR_CARTAO, fg=COR_TEXTO_2
    ).grid(row=0, column=0, sticky="w", padx=(0, 16), pady=10)

    combo = criar_entry(form)
    combo.grid(row=0, column=1, sticky="ew", ipady=8, pady=10)

    tk.Label(
        form, text="Apartamento de destino *", font=(FONTE, 10, "bold"),
        bg=COR_CARTAO, fg=COR_TEXTO_2
    ).grid(row=2, column=0, sticky="w", padx=(0, 16), pady=10)

    apto = criar_entry(form)
    apto.grid(row=2, column=1, sticky="ew", ipady=8, pady=10)

    mapa = {}
    nomes = []

    try:
        for prestador in listar_prestadores():
            texto = f"{prestador[1]} | {prestador[2]} | {prestador[3]}"
            nomes.append(texto)
            mapa[texto] = prestador[0]
    except Exception as erro:
        messagebox.showerror(
            "Erro",
            f"Não foi possível carregar os prestadores.\n\n{erro}",
            parent=janela
        )

    # Lista de sugestões exibida logo abaixo do campo.
    lista_sugestoes = tk.Listbox(
        form,
        height=6,
        font=(FONTE, 10),
        bg="#FFFFFF",
        fg="#17202A",
        selectbackground="#D9E8F6",
        selectforeground="#17202A",
        relief="flat",
        bd=0,
        highlightthickness=1,
        highlightbackground=COR_BORDA,
        activestyle="none",
    )

    def esconder_sugestoes():
        lista_sugestoes.grid_remove()

    def mostrar_sugestoes(itens):
        lista_sugestoes.delete(0, tk.END)

        for item in itens[:8]:
            lista_sugestoes.insert(tk.END, item)

        if itens:
            lista_sugestoes.grid(
                row=1,
                column=1,
                sticky="ew",
                pady=(0, 4),
            )
        else:
            esconder_sugestoes()

    def filtrar_prestadores(_evento=None):
        termo = combo.get().strip().lower()

        if not termo:
            esconder_sugestoes()
            return

        filtrados = [
            item for item in nomes
            if termo in item.lower()
        ]

        mostrar_sugestoes(filtrados)

    def selecionar_sugestao(_evento=None):
        selecao = lista_sugestoes.curselection()

        if not selecao:
            return

        selecionado = lista_sugestoes.get(selecao[0])

        combo.delete(0, tk.END)
        combo.insert(0, selecionado)

        esconder_sugestoes()
        apto.focus_set()

    def selecionar_primeiro(_evento=None):
        if lista_sugestoes.size() > 0:
            lista_sugestoes.selection_clear(0, tk.END)
            lista_sugestoes.selection_set(0)
            selecionar_sugestao()
            return "break"

    def ir_para_lista(_evento=None):
        if lista_sugestoes.size() > 0:
            lista_sugestoes.focus_set()
            lista_sugestoes.selection_clear(0, tk.END)
            lista_sugestoes.selection_set(0)
            lista_sugestoes.activate(0)
            return "break"

    combo.bind("<KeyRelease>", filtrar_prestadores)
    combo.bind("<Return>", selecionar_primeiro)
    combo.bind("<Down>", ir_para_lista)

    lista_sugestoes.bind("<Double-1>", selecionar_sugestao)
    lista_sugestoes.bind("<Return>", selecionar_sugestao)
    lista_sugestoes.bind("<Escape>", lambda e: (esconder_sugestoes(), combo.focus_set()))

    esconder_sugestoes()

    aviso = (
        f"{len(nomes)} prestador(es) disponível(is) para seleção."
        if nomes
        else "Nenhum prestador cadastrado."
    )

    tk.Label(
        card, text=aviso, font=(FONTE, 9),
        bg=COR_CARTAO, fg=COR_TEXTO_2
    ).pack(anchor="w", padx=32, pady=(4, 8))

    def registrar():
        selecionado = combo.get()
        apartamento = apto.get().strip()

        if not selecionado:
            messagebox.showwarning("Atenção", "Selecione um prestador.", parent=janela)
            combo.focus_set()
            return

        if selecionado not in mapa:
            termo = selecionado.strip().lower()
            correspondencias = [
                item for item in nomes
                if termo in item.lower()
            ]

            if len(correspondencias) == 1:
                selecionado = correspondencias[0]
                combo.delete(0, tk.END)
                combo.insert(0, selecionado)
            else:
                messagebox.showwarning(
                    "Atenção",
                    "Digite parte do nome e selecione um prestador da lista.",
                    parent=janela
                )
                combo.focus_set()
                return

        if not apartamento:
            messagebox.showwarning(
                "Atenção",
                "Informe o apartamento de destino.",
                parent=janela
            )
            apto.focus_set()
            return

        agora = datetime.now()
        data = agora.strftime("%d/%m/%Y")
        hora = agora.strftime("%H:%M:%S")

        try:
            registrar_entrada_prestador(
                mapa[selecionado],
                apartamento,
                data,
                hora,
                operador_atual["id"]
            )
        except Exception as erro:
            messagebox.showerror(
                "Erro",
                f"Não foi possível registrar a entrada.\n\n{erro}",
                parent=janela
            )
            return

        messagebox.showinfo(
            "Entrada registrada",
            f"Entrada registrada com sucesso!\n\n"
            f"Data: {data}\n"
            f"Hora: {hora}\n"
            f"Operador: {operador_atual['nome']}",
            parent=janela
        )

        combo.delete(0, tk.END)
        esconder_sugestoes()
        apto.delete(0, tk.END)
        combo.focus_set()

    botoes = tk.Frame(card, bg=COR_CARTAO)
    botoes.pack(fill="x", padx=32, pady=(16, 28))

    criar_botao(
        botoes, "Registrar entrada", registrar,
        COR_SUCESSO, COR_SUCESSO_HOVER
    ).pack(side="right")

    apto.bind("<Return>", lambda e: registrar())
    combo.focus_set()


def abrir_acessos_ativos(janela_principal, operador_atual):
    janela = tk.Toplevel(janela_principal)
    janela.title("Acessos Ativos")
    janela.configure(bg=COR_FUNDO)
    janela.minsize(1050, 620)
    centralizar(janela, 1220, 720)
    janela.transient(janela_principal)

    configurar_tabela("Acessos.Treeview")
    criar_topo(
        janela,
        "Acessos Ativos",
        "Prestadores que ainda possuem entrada ativa no condomínio."
    )

    conteudo = tk.Frame(janela, bg=COR_FUNDO)
    conteudo.pack(fill="both", expand=True, padx=30, pady=24)

    faixa = tk.Frame(
        conteudo, bg=COR_CARTAO,
        highlightthickness=1, highlightbackground=COR_BORDA
    )
    faixa.pack(fill="x", pady=(0, 16))

    total = tk.Label(
        faixa, text="0 acessos ativos", font=(FONTE, 11, "bold"),
        bg=COR_CARTAO, fg=COR_TEXTO
    )
    total.pack(side="left", padx=18, pady=14)

    quadro = tk.Frame(
        conteudo, bg=COR_CARTAO,
        highlightthickness=1, highlightbackground=COR_BORDA
    )
    quadro.pack(fill="both", expand=True)
    quadro.grid_rowconfigure(0, weight=1)
    quadro.grid_columnconfigure(0, weight=1)

    colunas = ("id", "nome", "documento", "empresa", "apartamento", "data", "hora")
    tabela = ttk.Treeview(
        quadro, columns=colunas, show="headings",
        style="Acessos.Treeview", selectmode="browse"
    )

    titulos = ("ID", "Nome", "Documento", "Empresa", "Apartamento", "Data Entrada", "Hora Entrada")
    for coluna, titulo in zip(colunas, titulos):
        tabela.heading(coluna, text=titulo)

    tabela.column("id", width=60, anchor="center", stretch=False)
    tabela.column("nome", width=260)
    tabela.column("documento", width=170)
    tabela.column("empresa", width=230)
    tabela.column("apartamento", width=120, anchor="center")
    tabela.column("data", width=125, anchor="center")
    tabela.column("hora", width=115, anchor="center")

    sv = ttk.Scrollbar(quadro, orient="vertical", command=tabela.yview)
    sh = ttk.Scrollbar(quadro, orient="horizontal", command=tabela.xview)
    tabela.configure(yscrollcommand=sv.set, xscrollcommand=sh.set)

    tabela.grid(row=0, column=0, sticky="nsew")
    sv.grid(row=0, column=1, sticky="ns")
    sh.grid(row=1, column=0, sticky="ew")

    tabela.tag_configure("par", background=COR_TABELA)
    tabela.tag_configure("impar", background=COR_TABELA_ALT)

    def carregar():
        for item in tabela.get_children():
            tabela.delete(item)

        try:
            acessos = list(listar_acessos_ativos())
        except Exception as erro:
            messagebox.showerror(
                "Erro",
                f"Não foi possível carregar os acessos.\n\n{erro}",
                parent=janela
            )
            return

        for i, acesso in enumerate(acessos):
            tabela.insert(
                "", tk.END, values=acesso,
                tags=("par" if i % 2 == 0 else "impar",)
            )

        total.config(text=f"{len(acessos)} acesso(s) ativo(s)")

    def registrar_saida():
        selecionado = tabela.selection()

        if not selecionado:
            messagebox.showwarning(
                "Atenção",
                "Selecione um prestador.",
                parent=janela
            )
            return

        dados = tabela.item(selecionado[0])["values"]

        confirmar = messagebox.askyesno(
            "Confirmar saída",
            f"Registrar saída deste prestador?\n\n"
            f"Nome: {dados[1]}\nApartamento: {dados[4]}",
            parent=janela
        )

        if not confirmar:
            return

        agora = datetime.now()
        data = agora.strftime("%d/%m/%Y")
        hora = agora.strftime("%H:%M:%S")

        try:
            registrar_saida_prestador(
                dados[0],
                data,
                hora,
                operador_atual["id"]
            )
        except Exception as erro:
            messagebox.showerror(
                "Erro",
                f"Não foi possível registrar a saída.\n\n{erro}",
                parent=janela
            )
            return

        messagebox.showinfo(
            "Saída registrada",
            f"Saída registrada com sucesso!\n\n"
            f"Data: {data}\n"
            f"Hora: {hora}\n"
            f"Operador: {operador_atual['nome']}",
            parent=janela
        )
        carregar()

    acoes = tk.Frame(conteudo, bg=COR_FUNDO)
    acoes.pack(fill="x", pady=(16, 0))

    criar_botao(
        acoes, "Registrar saída", registrar_saida,
        COR_PERIGO, COR_PERIGO_HOVER
    ).pack(side="right")

    criar_botao(
        acoes, "Atualizar lista", carregar
    ).pack(side="right", padx=(0, 10))

    tabela.bind("<Double-1>", lambda e: registrar_saida())
    janela.bind("<F5>", lambda e: carregar())

    carregar()


def abrir_historico_acessos(janela_principal):
    janela = tk.Toplevel(janela_principal)
    janela.title("Histórico de Acessos")
    janela.configure(bg=COR_FUNDO)
    janela.minsize(1150, 650)
    centralizar(janela, 1380, 760)
    janela.transient(janela_principal)

    configurar_tabela("Historico.Treeview")
    criar_topo(
        janela,
        "Histórico de Acessos",
        "Consulte entradas e saídas de prestadores."
    )

    conteudo = tk.Frame(janela, bg=COR_FUNDO)
    conteudo.pack(fill="both", expand=True, padx=30, pady=22)

    filtros = tk.Frame(
        conteudo, bg=COR_CARTAO,
        highlightthickness=1, highlightbackground=COR_BORDA
    )
    filtros.pack(fill="x", pady=(0, 16))

    tk.Label(
        filtros, text="Nome", font=(FONTE, 9, "bold"),
        bg=COR_CARTAO, fg=COR_TEXTO_2
    ).grid(row=0, column=0, padx=(18, 6), pady=14)

    nome = criar_entry(filtros, 22)
    nome.grid(row=0, column=1, ipady=7, padx=(0, 16), pady=11)

    tk.Label(
        filtros, text="Apartamento", font=(FONTE, 9, "bold"),
        bg=COR_CARTAO, fg=COR_TEXTO_2
    ).grid(row=0, column=2, padx=(0, 6))

    apto = criar_entry(filtros, 12)
    apto.grid(row=0, column=3, ipady=7, padx=(0, 16))

    tk.Label(
        filtros, text="Data", font=(FONTE, 9, "bold"),
        bg=COR_CARTAO, fg=COR_TEXTO_2
    ).grid(row=0, column=4, padx=(0, 6))

    data = criar_entry(filtros, 12)
    data.grid(row=0, column=5, ipady=7, padx=(0, 8))

    quadro = tk.Frame(
        conteudo, bg=COR_CARTAO,
        highlightthickness=1, highlightbackground=COR_BORDA
    )
    quadro.pack(fill="both", expand=True)
    quadro.grid_rowconfigure(0, weight=1)
    quadro.grid_columnconfigure(0, weight=1)

    colunas = (
        "id", "nome", "documento", "empresa", "apartamento",
        "data_entrada", "hora_entrada", "data_saida", "hora_saida", "status",
        "operador_entrada", "operador_saida"
    )
    tabela = ttk.Treeview(
        quadro, columns=colunas, show="headings",
        style="Historico.Treeview", selectmode="browse"
    )

    titulos = (
        "ID", "Nome", "Documento", "Empresa", "Apartamento",
        "Data Entrada", "Hora Entrada", "Data Saída", "Hora Saída", "Status",
        "Operador Entrada", "Operador Saída"
    )

    for coluna, titulo in zip(colunas, titulos):
        tabela.heading(coluna, text=titulo)

    larguras = (55, 220, 150, 190, 105, 110, 100, 110, 100, 100, 180, 180)
    for coluna, largura in zip(colunas, larguras):
        tabela.column(
            coluna, width=largura,
            anchor="center" if coluna not in ("nome", "documento", "empresa") else "w"
        )

    sv = ttk.Scrollbar(quadro, orient="vertical", command=tabela.yview)
    sh = ttk.Scrollbar(quadro, orient="horizontal", command=tabela.xview)
    tabela.configure(yscrollcommand=sv.set, xscrollcommand=sh.set)

    tabela.grid(row=0, column=0, sticky="nsew")
    sv.grid(row=0, column=1, sticky="ns")
    sh.grid(row=1, column=0, sticky="ew")

    tabela.tag_configure("par", background=COR_TABELA)
    tabela.tag_configure("impar", background=COR_TABELA_ALT)

    def preencher(registros):
        for item in tabela.get_children():
            tabela.delete(item)

        registros = list(registros)
        for i, registro in enumerate(registros):
            tabela.insert(
                "", tk.END, values=registro,
                tags=("par" if i % 2 == 0 else "impar",)
            )

    def carregar():
        try:
            preencher(listar_historico_acessos())
        except Exception as erro:
            messagebox.showerror(
                "Erro",
                f"Não foi possível carregar o histórico.\n\n{erro}",
                parent=janela
            )

    def pesquisar():
        nome_pesquisa = nome.get().strip().lower()
        apto_pesquisa = apto.get().strip().lower()
        data_pesquisa = data.get().strip().lower()

        try:
            filtrados = []

            for registro in listar_historico_acessos():
                ok_nome = not nome_pesquisa or nome_pesquisa in str(registro[1]).lower()
                ok_apto = not apto_pesquisa or apto_pesquisa in str(registro[4]).lower()
                ok_data = (
                    not data_pesquisa
                    or data_pesquisa in str(registro[5]).lower()
                    or data_pesquisa in str(registro[7]).lower()
                )

                if ok_nome and ok_apto and ok_data:
                    filtrados.append(registro)

            preencher(filtrados)

        except Exception as erro:
            messagebox.showerror(
                "Erro",
                f"Não foi possível pesquisar o histórico.\n\n{erro}",
                parent=janela
            )

    def limpar():
        nome.delete(0, tk.END)
        apto.delete(0, tk.END)
        data.delete(0, tk.END)
        carregar()
        nome.focus_set()

    def exportar_excel():
        itens = tabela.get_children()

        if not itens:
            messagebox.showwarning(
                "Exportar Excel",
                "Não há registros na tabela para exportar.",
                parent=janela,
            )
            return

        nome_padrao = (
            "historico_acessos_"
            + datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            + ".xlsx"
        )

        caminho = filedialog.asksaveasfilename(
            parent=janela,
            title="Salvar histórico de acessos",
            defaultextension=".xlsx",
            initialfile=nome_padrao,
            filetypes=[("Planilha do Excel", "*.xlsx")],
        )

        if not caminho:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Histórico de Acessos"

            cabecalhos = [
                "ID",
                "Nome",
                "Documento",
                "Empresa",
                "Apartamento",
                "Data Entrada",
                "Hora Entrada",
                "Data Saída",
                "Hora Saída",
                "Status",
                "Operador Entrada",
                "Operador Saída",
            ]

            # Título
            ws.merge_cells("A1:L1")
            titulo = ws["A1"]
            titulo.value = "Histórico de Acessos de Prestadores"
            titulo.font = Font(bold=True, size=16, color="FFFFFF")
            titulo.fill = PatternFill("solid", fgColor="111D30")
            titulo.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 26

            # Informações do filtro usado
            filtros_usados = []
            if nome.get().strip():
                filtros_usados.append(f"Nome: {nome.get().strip()}")
            if apto.get().strip():
                filtros_usados.append(f"Apartamento: {apto.get().strip()}")
            if data.get().strip():
                filtros_usados.append(f"Data: {data.get().strip()}")

            ws.merge_cells("A2:L2")
            info = ws["A2"]
            info.value = (
                "Filtros: " + " | ".join(filtros_usados)
                if filtros_usados
                else "Filtros: nenhum — todos os registros exibidos"
            )
            info.font = Font(italic=True, size=10, color="5B6573")
            info.alignment = Alignment(horizontal="left")

            # Cabeçalho da tabela
            preenchimento_cabecalho = PatternFill("solid", fgColor="1B2D45")
            fonte_cabecalho = Font(bold=True, color="FFFFFF")
            borda = Border(
                bottom=Side(style="thin", color="C8D1DC")
            )

            for coluna, texto in enumerate(cabecalhos, start=1):
                celula = ws.cell(row=4, column=coluna, value=texto)
                celula.fill = preenchimento_cabecalho
                celula.font = fonte_cabecalho
                celula.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
                celula.border = borda

            # Exporta exatamente o que está visível na tabela
            for linha_excel, item_id in enumerate(itens, start=5):
                valores = tabela.item(item_id)["values"]

                for coluna, valor in enumerate(valores, start=1):
                    celula = ws.cell(
                        row=linha_excel,
                        column=coluna,
                        value=valor,
                    )
                    celula.alignment = Alignment(
                        vertical="center",
                        horizontal=(
                            "center"
                            if coluna in (1, 5, 6, 7, 8, 9, 10)
                            else "left"
                        ),
                    )

                if linha_excel % 2 == 0:
                    preenchimento = PatternFill(
                        "solid",
                        fgColor="F3F6FA",
                    )
                    for coluna in range(1, 13):
                        ws.cell(
                            row=linha_excel,
                            column=coluna,
                        ).fill = preenchimento

            # Larguras de coluna
            larguras = {
                "A": 8,
                "B": 28,
                "C": 20,
                "D": 26,
                "E": 14,
                "F": 14,
                "G": 13,
                "H": 14,
                "I": 13,
                "J": 14,
                "K": 24,
                "L": 24,
            }

            for coluna, largura in larguras.items():
                ws.column_dimensions[coluna].width = largura

            ws.freeze_panes = "A5"
            ws.auto_filter.ref = f"A4:L{4 + len(itens)}"

            wb.save(caminho)

            messagebox.showinfo(
                "Exportação concluída",
                "Histórico exportado com sucesso!\n\n"
                f"Arquivo:\n{caminho}",
                parent=janela,
            )

        except Exception as erro:
            messagebox.showerror(
                "Erro na exportação",
                "Não foi possível gerar a planilha.\n\n"
                f"{erro}",
                parent=janela,
            )

    criar_botao(
        filtros, "Pesquisar", pesquisar
    ).grid(row=0, column=6, padx=(8, 6))

    criar_botao(
        filtros, "Limpar filtros", limpar
    ).grid(row=0, column=7, padx=(0, 6))

    criar_botao(
        filtros,
        "Exportar Excel",
        exportar_excel,
        COR_SUCESSO,
        COR_SUCESSO_HOVER,
    ).grid(row=0, column=8, padx=(0, 14))

    for campo in (nome, apto, data):
        campo.bind("<Return>", lambda e: pesquisar())

    janela.bind("<F5>", lambda e: carregar())

    carregar()
    nome.focus_set()